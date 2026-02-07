import os
import pandas as pd
from django.core.management.base import BaseCommand
from django.conf import settings
from schedules.models import Teacher
from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'Carga maestros desde un archivo Excel (Secciones de los Maestros.xlsx) eliminando duplicados'

    def handle(self, *args, **options):
        # Ruta al archivo Excel (asumiendo que está en el mismo directorio que manage.py)
        file_path = os.path.join(settings.BASE_DIR, 'data', 'Secciones de los Maestros.xlsx')
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"El archivo {file_path} no existe!"))
            return

        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active
            
            # Extraer nombres de la columna A (sin cabecera si existe)
            teacher_names = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Asume que la primera fila es cabecera
                if row[0]:  # Si la celda A no está vacía
                    teacher_names.append(row[0].strip())
            
            # Eliminar duplicados manteniendo el orden
            unique_names = list(dict.fromkeys(teacher_names))
            
            # Contadores para estadísticas
            created = 0
            existing = 0
            
            for name in unique_names:
                # Verificar si el maestro ya existe (insensible a mayúsculas)
                if not Teacher.objects.filter(full_name__iexact=name).exists():
                    Teacher.objects.create(full_name=name)
                    created += 1
                else:
                    existing += 1
            
            self.stdout.write(self.style.SUCCESS(
                f"Proceso completado. Maestros nuevos: {created}, ya existentes: {existing}"
            ))
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Ocurrió un error: {str(e)}"))